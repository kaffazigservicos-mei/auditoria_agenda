from openpyxl import load_workbook
from pathlib import Path

path = Path('/home/ubuntu/auditoria_agenda/planilha_registro_atendimentos.xlsx')
wb = load_workbook(path, data_only=False)
ws = wb['Registros']
assert ws['D2'].value == 160
assert ws['B3'].value == '=SUM(E6:E110)'
assert ws['D3'].value == '=IF(B3=0,"",IF(B3<D2,"Déficit de "&(D2-B3)&" min",IF(B3>D2,"Crédito de "&(B3-D2)&" min","Sem crédito ou déficit")))'
assert ws['E6'].value == '=IF(OR(B6="",C6=""),"",ROUND((C6-B6)*1440,0))'
assert ws['E110'].value == '=IF(OR(B110="",C110=""),"",ROUND((C110-B110)*1440,0))'
assert len(ws.data_validations.dataValidation) == 2
assert any('Normal,Excepcional' in str(dv.formula1) for dv in ws.data_validations.dataValidation)
assert len(ws.conditional_formatting) >= 2
assert any('LEFT(D3,7)="Déficit"' in str(rule.formula) for rules in ws.conditional_formatting._cf_rules.values() for rule in rules)
assert ws.freeze_panes == 'A6'
assert ws['B2'].number_format == 'dd/mm/yyyy'
assert ws['B6'].number_format == 'hh:mm'
assert ws['C6'].number_format == 'hh:mm'
assert ws['E6'].number_format == '0'
assert ws['A6'].number_format == '@'
assert ws['D6'].number_format == '@'
assert wb['Resumo']['B4'].value == '=Registros!D2'
assert wb['Resumo']['B5'].value == '=Registros!B3'
assert wb['Resumo']['B6'].value == '=Registros!D3'
assert any('LEFT(B6,7)="Déficit"' in str(rule.formula) for rules in wb['Resumo'].conditional_formatting._cf_rules.values() for rule in rules)

def resultado(total, previsao=160):
    if total == 0:
        return ''
    if total < previsao:
        return f'Déficit de {previsao-total} min'
    if total > previsao:
        return f'Crédito de {total-previsao} min'
    return 'Sem crédito ou déficit'

assert resultado(140) == 'Déficit de 20 min'
assert resultado(160) == 'Sem crédito ou déficit'
assert resultado(181) == 'Crédito de 21 min'
print('Cenários testados: 140 -> Déficit de 20 min; 160 -> Sem crédito ou déficit; 181 -> Crédito de 21 min.')
print('VALIDAÇÃO OK')
print('Abas:', wb.sheetnames)
print('Linhas de registro:', 105)
print('Fórmulas e formatações essenciais conferidas.')
