from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment

OUTPUT = "/home/ubuntu/auditoria_agenda/planilha_registro_atendimentos.xlsx"
MAX_ROWS = 105
FIRST_DATA_ROW = 6
LAST_DATA_ROW = FIRST_DATA_ROW + MAX_ROWS - 1

wb = Workbook()
ws = wb.active
ws.title = "Registros"
resumo = wb.create_sheet("Resumo")

# Paleta de alto contraste para leitura no celular
blue = "1F4E78"
light_blue = "D9EAF7"
green = "E2F0D9"
yellow = "FFF2CC"
red_fill = "F4CCCC"
red_font = "9C0006"
gray = "F3F6F9"
white = "FFFFFF"
dark = "17365D"
thin_gray = Side(style="thin", color="D9E2F3")

# Configuração da aba principal
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A6"
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.outlinePr.summaryBelow = True

ws.merge_cells("A1:E1")
ws["A1"] = "REGISTRO DE ATENDIMENTOS ONLINE"
ws["A1"].font = Font(bold=True, color=white, size=14)
ws["A1"].fill = PatternFill("solid", fgColor=blue)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Campos de cabeçalho / parâmetros
labels = [("A2", "Data do atendimento"), ("C2", "Previsão total (minutos)"), ("A3", "Soma de minutos"), ("C3", "Saldo / déficit (minutos)")]
for cell, value in labels:
    ws[cell] = value
    ws[cell].font = Font(bold=True, color=dark)
    ws[cell].fill = PatternFill("solid", fgColor=light_blue)
    ws[cell].alignment = Alignment(wrap_text=True, vertical="center")

ws["B2"] = ""
ws["B2"].number_format = "dd/mm/yyyy"
ws["B2"].fill = PatternFill("solid", fgColor=yellow)
ws["B2"].comment = Comment("Informe a data referente aos atendimentos registrados.", "Manus AI")
ws["D2"] = 160
ws["D2"].number_format = "0"
ws["D2"].fill = PatternFill("solid", fgColor=yellow)
ws["D2"].comment = Comment("Previsão padrão definida em 160 minutos. Pode ser alterada se necessário.", "Manus AI")
ws["B3"] = f"=SUM(E{FIRST_DATA_ROW}:E{LAST_DATA_ROW})"
ws["B3"].number_format = "0"
ws["B3"].fill = PatternFill("solid", fgColor=green)
ws["D3"] = "=D2-B3"
ws["D3"].number_format = "0;[Red]-0;0"
ws["D3"].fill = PatternFill("solid", fgColor=green)
for cell in ["B2", "D2", "B3", "D3"]:
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].font = Font(bold=True, color=dark, size=11)

# Cabeçalho da tabela
headers = ["Classificação", "Hora inicial", "Hora encerramento", "Nome", "Total de atendimento (min)"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(FIRST_DATA_ROW - 1, col, header)
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=dark)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(top=thin_gray, bottom=thin_gray)
ws.row_dimensions[FIRST_DATA_ROW - 1].height = 36

# Linhas de registro com fórmula automática
for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
    for col in range(1, 6):
        cell = ws.cell(row, col)
        cell.border = Border(bottom=thin_gray)
        cell.alignment = Alignment(vertical="center", wrap_text=(col in [1, 4]))
        if row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=gray)
    # Campos de texto: força teclado alfabético em aplicativos móveis compatíveis.
    ws.cell(row, 1).number_format = "@"
    ws.cell(row, 4).number_format = "@"
    ws.cell(row, 2).number_format = "hh:mm"
    ws.cell(row, 3).number_format = "hh:mm"
    ws.cell(row, 5).value = f'=IF(OR(B{row}="",C{row}=""),"",ROUND((C{row}-B{row})*1440,0))'
    ws.cell(row, 5).number_format = "0"
    ws.cell(row, 5).font = Font(color=dark)

# Tabela com filtro, útil no celular
ref = f"A{FIRST_DATA_ROW-1}:E{LAST_DATA_ROW}"
tab = Table(displayName="TabelaAtendimentos", ref=ref)
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)

# Validação da classificação
validation = DataValidation(type="list", formula1='"Normal,Excepcional"', allow_blank=True)
validation.error = "Escolha Normal ou Excepcional."
validation.errorTitle = "Classificação inválida"
validation.prompt = "Selecione o tipo de atendimento."
validation.promptTitle = "Classificação"
ws.add_data_validation(validation)
validation.add(f"A{FIRST_DATA_ROW}:A{LAST_DATA_ROW}")

# Validação para horários
hora_validation = DataValidation(type="time", operator="between", formula1="0", formula2="1", allow_blank=True)
hora_validation.error = "Informe um horário válido, por exemplo 08:30."
hora_validation.errorTitle = "Horário inválido"
ws.add_data_validation(hora_validation)
hora_validation.add(f"B{FIRST_DATA_ROW}:C{LAST_DATA_ROW}")

# Formatação condicional: déficit geral em vermelho; linhas excepcionais em amarelo suave
ws.conditional_formatting.add("D3", FormulaRule(formula=["D3>0"], fill=PatternFill("solid", fgColor=red_fill), font=Font(color=red_font, bold=True)))
ws.conditional_formatting.add("B3", FormulaRule(formula=["B3<D2"], fill=PatternFill("solid", fgColor=red_fill), font=Font(color=red_font, bold=True)))
ws.conditional_formatting.add(f"A{FIRST_DATA_ROW}:E{LAST_DATA_ROW}", FormulaRule(formula=[f'$A{FIRST_DATA_ROW}="Excepcional"'], fill=PatternFill("solid", fgColor=yellow)))

# Dimensões e acessibilidade no celular
widths = {"A": 17, "B": 13, "C": 16, "D": 28, "E": 24}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
    ws.row_dimensions[row].height = 24

ws.auto_filter.ref = ref
ws.sheet_properties.tabColor = blue

# Aba de resumo com explicação de uso
resumo.sheet_view.showGridLines = False
resumo.merge_cells("A1:D1")
resumo["A1"] = "RESUMO E INSTRUÇÕES"
resumo["A1"].font = Font(bold=True, color=white, size=14)
resumo["A1"].fill = PatternFill("solid", fgColor=blue)
resumo["A1"].alignment = Alignment(horizontal="center")
resumo.row_dimensions[1].height = 28
resumo["A3"] = "Indicador"
resumo["B3"] = "Valor"
for cell in ["A3", "B3"]:
    resumo[cell].font = Font(bold=True, color=white)
    resumo[cell].fill = PatternFill("solid", fgColor=dark)
resumo["A4"] = "Previsão de atendimento (minutos)"
resumo["B4"] = "=Registros!D2"
resumo["A5"] = "Soma de minutos registrados"
resumo["B5"] = "=Registros!B3"
resumo["A6"] = "Saldo / déficit"
resumo["B6"] = "=Registros!D3"
for row in range(4, 7):
    resumo[f"A{row}"].fill = PatternFill("solid", fgColor=light_blue)
    resumo[f"B{row}"].fill = PatternFill("solid", fgColor=green)
    resumo[f"A{row}"].alignment = resumo[f"B{row}"].alignment = Alignment(vertical="center", wrap_text=True)
    resumo[f"B{row}"].font = Font(bold=True, color=dark)
resumo["B6"].number_format = "0;[Red]-0;0"
resumo.conditional_formatting.add("B6", FormulaRule(formula=["B6>0"], fill=PatternFill("solid", fgColor=red_fill), font=Font(color=red_font, bold=True)))
resumo.conditional_formatting.add("B5", FormulaRule(formula=["B5<B4"], fill=PatternFill("solid", fgColor=red_fill), font=Font(color=red_font, bold=True)))

resumo["A9"] = "Como usar"
resumo["A9"].font = Font(bold=True, color=white)
resumo["A9"].fill = PatternFill("solid", fgColor=dark)
resumo.merge_cells("A10:D10")
resumo["A10"] = "1. Na aba Registros, informe a data do atendimento e confirme a previsão de 160 minutos."
resumo.merge_cells("A11:D11")
resumo["A11"] = "2. Em cada linha, escolha Normal ou Excepcional, informe hora inicial, hora de encerramento e nome."
resumo.merge_cells("A12:D12")
resumo["A12"] = "3. O total de minutos de cada atendimento e a soma geral são calculados automaticamente."
resumo.merge_cells("A13:D13")
resumo["A13"] = "4. O saldo é calculado por: previsão de atendimentos − soma de minutos. Saldo negativo aparece em vermelho."
for row in range(10, 14):
    resumo[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
    resumo.row_dimensions[row].height = 32

for col, width in {"A": 42, "B": 22, "C": 18, "D": 18}.items():
    resumo.column_dimensions[col].width = width
resumo.sheet_properties.tabColor = "70AD47"

# Ativa a aba de registros ao abrir
wb.active = 0
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(OUTPUT)
print(OUTPUT)
